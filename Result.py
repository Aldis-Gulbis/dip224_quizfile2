from openpyxl import Workbook, load_workbook
import statistics

wb = load_workbook("EmployeeData.xlsx")
page = wb.active
rindas = page.max_row

a = 0
saraksts = []
vid_vecums = []

for i in range(2, rindas + 1):
    nodala = page["C" + str(i)].value
    if nodala == "Sales":
        a = 0 + 1
        vecums1 = page["E" + str (i)].value
        vid_vecums.append(vecums1)
        saraksts.append(a)

x = statistics.mean(vid_vecums)

print("3. ", int(round(x)))
print("1. ", len(saraksts))

min_vecums = []

for i in range(2, rindas + 1):
    nodala = page["C" + str(i)].value
    if nodala == "Finance":
        vecums2 = page["E" + str(i)].value
        min_vecums.append(vecums2)

print("2. ", min(min_vecums))

alga = []
for i in range(2, rindas + 1):
    annual_salary = page["F" + str(i)].value
    bonus = page["G" + str(i)].value
    salary = float(annual_salary) + (float(bonus) * float(annual_salary)) 
    if salary > 150000 and salary < 200000:
        alga.append(salary)

print("4. ",len(alga))

for i in range(2, rindas + 1):
    annual_salary = page["F" + str(i)].value
    bonus = page["G" + str(i)].value
    salary = float(annual_salary) + (float(bonus) * float(annual_salary))
    if salary == max(salary):
        x
wb.close()